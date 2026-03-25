from typing import Literal
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def sort_correlations(tickers_correlations: dict, sort_order: Literal["asc", "desc"]):
    return dict(
        sorted(
            tickers_correlations.items(),
            key=lambda item: item[1],
            reverse=(sort_order == "desc")
        )
    )
    
def filter_low_correlations(tickers_correlations: dict, threshold: float):
    return {
        ticker: corr 
        for ticker, corr in tickers_correlations.items() 
        if corr <= threshold
    }

def save_txt(tickers_correlations: dict, file_path: str) -> None:
    with open(f"{file_path}.txt", "w", encoding="utf-8") as f:
        for ticker, corr in tickers_correlations.items():
            f.write(f"{ticker}: {corr}\n")

def save_excel(tickers_correlations: dict, file_path: str, threshold: float) -> None:
    wb = Workbook()
    ws = wb.active

    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

    for ticker, corr in tickers_correlations.items():
        ws.append([ticker, corr])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.fill = green if corr <= threshold else red

    wb.save(f"{file_path}.xlsx")