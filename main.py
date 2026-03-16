from botasaurus.browser import browser, Driver
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn
from InquirerPy import inquirer
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import requests
import json

console = Console()

class TradingViewCorrs:
    def __init__(self, driver: Driver):
        self.driver = driver
        self.cookies = {}
        self.all_tickers = []
        self.tickers_correlations = {}
        self.session = requests.Session()
        retry = Retry(
            total=5,                       
            backoff_factor=1,          
            status_forcelist=[429,500,502,503,504],
            allowed_methods=["GET","POST"] 
        )
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    # -----------------------
    # Открыть TradingView
    # -----------------------
    def open_tradingview(self):
        with console.status("[green]Открываем TradingView...[/green]"):
            self.driver.get("https://ru.tradingview.com/chart/RRGuoDgP")
            self.cookies = self.driver.get_cookies_dict()
            self.session.cookies.update(self.cookies)
        console.print("[bold green]✔ Открыт TradingView[/bold green]")

    # -----------------------
    # Добавить тикеры в список
    # -----------------------
    def add_tickers_to_list(self, tickers):

        watchlist_get_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red?source=web"
        watchlist_add_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red/append/?source=web"
        watchlist_remove_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red/remove/?source=web"

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://ru.tradingview.com",
            "Referer": "https://ru.tradingview.com/chart/RRGuoDgP"
        }

        symbols_data = [f"BYBIT:{ticker}" for ticker in tickers]

        watchlist_get_response = self.session.get(url=watchlist_get_url, headers=headers)
        watchlist_get_response.raise_for_status()

        watchlist = watchlist_get_response.json()["symbols"]
        if watchlist:
            watchlist_remove_response = self.session.post(url=watchlist_remove_url, headers=headers, json=watchlist)
            watchlist_remove_response.raise_for_status()

        watchlist_add_response = self.session.post(url=watchlist_add_url, headers=headers, json=symbols_data)
        watchlist_add_response.raise_for_status()

    # -----------------------
    # Добавить индикатор корреляции
    # -----------------------
    def activate_corr_indicator(self):
        with console.status("[yellow]Активируем индикатор корреляции...[/yellow]", spinner="line"):
            self.driver.enable_human_mode()

            indicators = self.driver.select_all(".sources-l31H9iuA .mainTitle-l31H9iuA")
            corr_exists = False

            for indicator in indicators:
                if indicator.text == "Корреляция":
                    corr_exists = True
                    continue

                self.driver.move_mouse_to_element(indicator)
                indicator.parent.parent.click("button[data-qa-id='legend-delete-action']")

            self.driver.disable_human_mode()

            if not corr_exists:
                self.driver.click("#header-toolbar-indicators button[data-name='show-favorite-indicators']")
                self.driver.get_element_with_exact_text("Корреляция").click()
        console.print("[bold green]✔ Индикатор активирован[/bold green]")
    
    # -----------------------
    # Получить все тикеры
    # -----------------------
    def get_all_tickers(self):

        all_tickers_get_url = "https://scanner.tradingview.com/crypto/scan?label-product=popup-screener-crypto-cex"

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://www.tradingview.com",
            "Referer": "https://www.tradingview.com/"
        }

        with open("scanner_payload.json", encoding="utf-8") as f:
            payload = json.load(f)

        with console.status("[cyan]Получаем список всех тикеров...[/cyan]", spinner="bouncingBar"):
            response = self.session.post(url=all_tickers_get_url, headers=headers, json=payload)
            response.raise_for_status()
            data_response = response.json()
            self.all_tickers = [item["s"] for item in data_response["data"]]
        console.print("[bold green]✔ Получены тикеры[/bold green]")

    # -----------------------
    # Получить корреляцию тикера
    # -----------------------
    def get_correlation(self):
        ticker = self.driver.select("[data-qa-id='details-element symbol']").text
        correlation = float(
            self.driver.select(".sources-l31H9iuA .valueValue-l31H9iuA").text.replace(",", ".").replace("−", "-")
        )
        self.tickers_correlations[ticker] = correlation
    
    # -----------------------
    # Сбор всех корреляций
    # -----------------------
    def collect_correlations(self):
        with console.status("[yellow]Подготавливаем тикеры...[/yellow]"):
            self.driver.click("button[data-name='screener-dialog-button']")
            
            template_name_screener = self.driver.select("div[data-qa-id='screen-title']")
            if template_name_screener.text != "Фьючерсы Bybit":
                template_name_screener.click()
                self.driver.get_element_with_exact_text("Фьючерсы Bybit").click()

            total_tickers = len(self.all_tickers)
            tickers = self.driver.select_all("tbody tr")
        console.print("[bold green]✔ Список тикеров готов[/bold green]")

        console.print(f"[magenta]Найдено {total_tickers} тикеров...[/magenta]")

        progress = Progress(
            SpinnerColumn(),
            TextColumn("[cyan]{task.description}"),
            BarColumn(complete_style="green", finished_style="bright_green"),
            TextColumn("[bright_green]{task.completed}[/bright_green]/[yellow]{task.total}[/yellow]"),
            TextColumn("[bright_green]{task.percentage:>3.0f}%[/bright_green]"),
            TimeRemainingColumn()
        )
        
        with progress:
            task = progress.add_task("Собираем корреляции", total=total_tickers)
            i = 0
            while i < len(tickers):
                tickers[i].click()
                self.get_correlation()
                i += 1
                progress.update(task, advance=1)

                if i >= len(tickers) - 5:
                    tickers = self.driver.select_all("tbody tr")
                
                if len(self.tickers_correlations) == total_tickers:
                    break
        
        print()
        console.print(Panel.fit(
            "🎯 [bold white]Все корреляции собраны![/bold white] 🎯",
            border_style="bold green",
            style="on dark_green",
            title="[bold yellow]Завершено[/bold yellow]"
        ))
        print()
    
    # =======================
    # Вывод результатов
    # =======================

    # -----------------------
    # Получение настроек 
    # -----------------------
    def ask_user_settings(self):
        threshold = inquirer.select(
            message="Выберите порог корреляции:",
            choices=[f"{x/10:.1f}" for x in range(1, 11)],
            default="0.5",
            qmark="", amark=""
        ).execute()
        threshold = float(threshold)

        sort_order = inquirer.select(
            message="Сортировка:",
            choices=[
                {"name": "без сортировки", "value": None},
                {"name": "сначала низкая корреляция", "value": "asc"},
                {"name": "сначала высокая корреляция", "value": "desc"},
            ],
            default=None,
            qmark="", amark=""
        ).execute()

        return threshold, sort_order

    # -----------------------
    # Сортировка корреляций
    # -----------------------
    def sort_correlations(self, sort_order):
        if sort_order:
            self.tickers_correlations = dict(
                sorted(
                    self.tickers_correlations.items(),
                    key=lambda item: item[1],
                    reverse=(sort_order == "desc")
                )
            )

    # -----------------------
    # Сохранение результатов в файлы
    # -----------------------
    def save_results_to_files(self, threshold):
        file_name = f"Корреляция_{datetime.now():%d.%m.%y_%H-%M}"
        
        with console.status("[green]Сохраняем результаты в файлы...[/green]"):
            # TXT
            with open(f"{file_name}.txt", "w", encoding="utf-8") as f:
                for ticker, corr in self.tickers_correlations.items():
                    f.write(f"{ticker}: {corr}\n")
            # Excel
            wb = Workbook()
            ws = wb.active
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
            for ticker, corr in self.tickers_correlations.items():
                ws.append([ticker, corr])
                cell = ws.cell(row=ws.max_row, column=2)
                cell.fill = green_fill if corr <= threshold else red_fill
            wb.save(f"{file_name}.xlsx")
        
        console.print(Panel.fit(
            f"🎯 [bold white]Результаты успешно сохранены![/bold white] 🎯\n\n"
            f"[cyan]Файлы:[/cyan]\n"
            f"• {file_name}.txt\n"
            f"• {file_name}.xlsx",
            border_style="bold green"
        ))
    
    # -----------------------
    # Показ таблицы в терминале
    # -----------------------
    def show_results_table(self, threshold):
        show_table = inquirer.select(
            message="Показать таблицу корреляций?",
            choices=[{"name": "Да", "value": True}, {"name": "Нет", "value": False}],
            default=True,
            qmark="", amark=""
        ).execute()

        if not show_table:
            return

        show_all_rows = inquirer.select(
            message=f"Все или только с корреляцией ниже {threshold}?",
            choices=[{"name": f"Ниже {threshold}", "value": False}, {"name": "Все", "value": True}],
            default=False,
            qmark="", amark=""
        ).execute()

        if not show_all_rows:
            data = {ticker: corr for ticker, corr in self.tickers_correlations.items() if corr <= threshold}
        else:
            data = self.tickers_correlations

        table = Table(title=f"\nКорреляции тикеров ({len(data)})", show_lines=True)
        table.add_column("Тикер", justify="center", style="cyan", no_wrap=True)
        table.add_column("Корреляция", justify="center")
        
        for ticker, corr in data.items():
            table.add_row(ticker, f"[green]{corr}[/green]")

        console.print(table)

    # -----------------------
    # Добавление тикеров с низкой корреляцией в список TradingView
    # -----------------------
    def add_low_corr_tickers_to_tradingview(self, threshold):
        add_to_watchlist = inquirer.select(
            message=f"Добавить тикеры с корреляцией ниже {threshold} в список TradingView?",
            choices=[
                {"name": "Да", "value": True}, 
                {"name": "Нет", "value": False}
            ],
            default=True,
            qmark="", amark=""
        ).execute()

        if not add_to_watchlist:
            return

        low_corr_tickers = [ticker for ticker, corr in self.tickers_correlations.items() if corr <= threshold]
        console.print(f"[cyan]Найдено {len(low_corr_tickers)} тикеров с корреляцией ниже {threshold}[/cyan]")

        batch_size = 30
        total = len(low_corr_tickers)

        for i in range(0, total, batch_size):
            ticker_batch = low_corr_tickers[i:i + batch_size]
            start = i + 1
            end = i + len(ticker_batch)

            with console.status(f"[cyan]Добавляем {start}-{end} из {len(low_corr_tickers)} тикеров в список TradingView...[/cyan]", spinner="line"):
                self.add_tickers_to_list(ticker_batch)
            
            if i + batch_size >= total:
                console.print(f"[bold green]Все {total} тикеров были успешно добавлены в список![/bold green]")
                break
            
            console.print(f"[yellow]Тикеры {start}-{end} добавлены в список.[/yellow]")

            action = inquirer.select(
                message="Продолжить добавление следующих тикеров?",
                choices=[
                    {"name": "Продолжить", "value": "continue"}, 
                    {"name": "Пропустить оставшиеся", "value": "skip"}
                ],
                default="continue",
                qmark="", amark=""
            ).execute()

            if action == "skip":
                console.print("[red]Добавление тикеров остановлено пользователем[/red]")
                break

        input("\nНажмите Enter для завершения...")
        self.session.close()

    # -----------------------
    # Основная функция display_results (только оркестрация)
    # -----------------------
    def display_results(self):
        console.print(f"[cyan]{'═'*25}[/cyan] [bold white]Вывод результатов[/bold white] [cyan]{'═'*25}[/cyan]")
        
        threshold, sort_order = self.ask_user_settings()
        self.sort_correlations(sort_order)
        self.save_results_to_files(threshold)
        self.show_results_table(threshold)
        self.add_low_corr_tickers_to_watchlist(threshold)

    # -----------------------
    # Pipeline: корреляции
    # -----------------------
    def get_correlations(self):
        self.open_tradingview()
        self.get_all_tickers()
        self.activate_corr_indicator()
        self.collect_correlations()
        self.display_results()

        return self.tickers_correlations


@browser(
    output=None,
    profile="TradingviewProfile",
    wait_for_complete_page_load=True
)
def main(driver: Driver, data):
    tv = TradingViewCorrs(driver)
    tickers_correlations = tv.get_correlations()

    
main()
