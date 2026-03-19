from botasaurus.browser import Driver
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from InquirerPy import inquirer
from contextlib import redirect_stdout
from PIL import Image
import requests
import json
import os
import time

console = Console()

class TradingViewCorrs:
    def __init__(self, driver: Driver):
        self.driver = driver
        self.cookies = {}
        self.all_tickers = []
        self.tickers_correlations = {}
        self.session = requests.Session()
        retry = Retry(
            total=5,                       
            backoff_factor=1,          
            status_forcelist=[429,500,502,503,504],
            allowed_methods=["GET","POST"] 
        )
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    # -----------------------
    # Открыть TradingView
    # -----------------------
    def open_tradingview(self):
        """Открывает TradingView"""
        with console.status("[green]Открываем TradingView...[/green]"):
            self.driver.get("https://ru.tradingview.com/chart/RRGuoDgP")

        if self.is_login_required():
            self.login_tradingview()

        self.cookies = self.driver.get_cookies_dict()
        self.session.cookies.update(self.cookies)

        console.print("[bold green]✔ TradingView готов к работе[/bold green]")

    def is_login_required(self):
        """Проверяет, требуется ли авторизация"""
        login_btn = self.driver.select(".linkButton-dfXNuaqf")
        return bool(login_btn)
    
    def login_tradingview(self):
        """Авторизация TradingView"""
        console.print("[bold bright_blue]Авторизация в аккаунт TradingView[/bold bright_blue]")

        while True:
            self.driver.click(".linkButton-dfXNuaqf")
            self.driver.click("button[name='Email']")

            username = input("Введите email: ")
            password = input("Введите пароль: ")

            with console.status("[cyan]Выполняем авторизацию...[/cyan]", spinner="dots"):
                self.driver.type("#id_username", username)
                self.driver.type("#id_password", password)
                self.driver.click("button[data-overflow-tooltip-text='Войти']")

            while True:  
                problem_elem = self.driver.select(".mainProblem-TCHLKPuQ")
                        
                if problem_elem is None:
                    with console.status("[cyan]Авторизация прошла успешно, загружаем TradingView...[/cyan]", spinner="dots"):
                        self.driver.reload()
                    return
                elif "Неправильное" in problem_elem.text:
                    console.print("[bright_red]Неверный логин или пароль. Попробуйте снова.[/bright_red]")
                    self.driver.reload()
                    break
                elif "CAPTCHA" in problem_elem.text:
                    if self.solve_captcha():
                        self.driver.click("button[data-overflow-tooltip-text='Войти']")
                        time.sleep(1.5)
                    else:
                        self.driver.reload()
                        break

    def solve_captcha(self):
        """Обработка reCAPTCHA через пользовательский ввод."""
        console.print(
            "[bright_yellow]" \
            "Вышла капча. Откройте изображение [underline]captcha_image.png[/underline] и введите номера нужных изображений через пробел.\n"
            "Может быть несколько капчей. В таком случае изображение обновится для повторного ввода номеров.\n"
            "Пустой ответ - если изображение прорисовано не полностью.\n"
            "0 - если изображение пустое.\n"
            "y - подтвердить капчу."
            "[/bright_yellow]"
        )

        captcha_frame = self.driver.select_iframe("iframe[title='reCAPTCHA']")
        captcha_frame.click("#rc-anchor-container")
        path_image = "output/screenshots/captcha_image.png"
        is_first_iteration = True

        while True:
            captcha_images = self.driver.select_iframe("[src*='bframe']").select("#rc-imageselect")
            self.get_captcha_image(path_image)
                        
            if not is_first_iteration:
                console.print("[bright_yellow]Изображение обновлено[/bright_yellow]")
            else:
                is_first_iteration = False

            user_input = input("Введите номера картинок: ")

            if user_input.lower() == "y":
                captcha_images.click(".rc-button-default")
                time.sleep(1.5)

                checkbox = captcha_frame.select("#recaptcha-anchor").get_attribute("aria-checked") == "true"
                if checkbox:
                    break
                else:
                    time.sleep(2)
                continue

            if user_input == "0":
                console.print("[bright_yellow]Перезагрузка авторизации...[/bright_yellow]")
                return False
            
            if user_input:
                self.click_captcha_images(captcha_images, user_input)
                time.sleep(1)
                continue

        console.print("[bold green]✔ Капча пройдена[/bold green]")
        return True

    
    def get_captcha_image(self, path_image):
        """Сохраняет скриншот и обрезает его по расположению CAPTCHA."""
        self.save_screenshot_silent(path_image)
        rect_captcha = self.driver.select("div:has(iframe[src*='bframe'])").get_bounding_rect()
        image = Image.open(path_image)
        cropped = image.crop((
            rect_captcha["x"],
            rect_captcha["y"],
            rect_captcha["x"] + rect_captcha["width"],
            rect_captcha["y"] + rect_captcha["height"]
        ))
        cropped.save(path_image)

    def save_screenshot_silent(self, filename):
        """Сохраняет скриншот без вывода в терминал"""
        with open(os.devnull, "w") as f, redirect_stdout(f):
            self.driver.save_screenshot(filename)

    def click_captcha_images(self, captcha_images, user_input):
        """Кликает по выбранным картинкам в CAPTCHA."""
        numbers = user_input.split(" ")
        try:
            for number in numbers:
                captcha_images.click(f"[id='{int(number) - 1}']")
                time.sleep(0.5)
        except ValueError:
            console.print("[red]Неккоректный ввод, попробуйте снова.[/red]")

    # -----------------------
    # Добавить тикеры в список
    # -----------------------
    def add_tickers_to_list(self, tickers):

        watchlist_get_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red?source=web"
        watchlist_add_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red/append/?source=web"
        watchlist_remove_url = "https://ru.tradingview.com/api/v1/symbols_list/colored/red/remove/?source=web"

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://ru.tradingview.com",
            "Referer": "https://ru.tradingview.com/chart/RRGuoDgP"
        }

        symbols_data = [f"BYBIT:{ticker}" for ticker in tickers]

        watchlist_get_response = self.session.get(url=watchlist_get_url, headers=headers)
        watchlist_get_response.raise_for_status()

        watchlist = watchlist_get_response.json()["symbols"]
        if watchlist:
            watchlist_remove_response = self.session.post(url=watchlist_remove_url, headers=headers, json=watchlist)
            watchlist_remove_response.raise_for_status()

        watchlist_add_response = self.session.post(url=watchlist_add_url, headers=headers, json=symbols_data)
        watchlist_add_response.raise_for_status()

    # -----------------------
    # Добавить индикатор корреляции
    # -----------------------
    def activate_corr_indicator(self):
        with console.status("[yellow]Активируем индикатор корреляции...[/yellow]", spinner="line"):
            self.driver.enable_human_mode()

            selector = ".sources-l31H9iuA .mainTitle-l31H9iuA"
            indicators = self.driver.select_all(selector)
            corr_exists = False

            for indicator in indicators:
                if indicator.text == "Корреляция":
                    corr_exists = True
                    continue
                
                self.driver.move_mouse_to_element(selector)
                indicator.parent.parent.click("button[data-qa-id='legend-delete-action']")

            self.driver.disable_human_mode()

            if not corr_exists:
                self.driver.click("#header-toolbar-indicators button[data-name='show-favorite-indicators']")
                self.driver.get_element_with_exact_text("Корреляция").click()
        console.print("[bold green]✔ Индикатор активирован[/bold green]")
    
    # -----------------------
    # Получить все тикеры
    # -----------------------
    def get_all_tickers(self):

        all_tickers_get_url = "https://scanner.tradingview.com/crypto/scan?label-product=popup-screener-crypto-cex"

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://www.tradingview.com",
            "Referer": "https://www.tradingview.com/"
        }

        with open("scanner_payload.json", encoding="utf-8") as f:
            payload = json.load(f)

        with console.status("[cyan]Получаем список всех тикеров...[/cyan]", spinner="bouncingBar"):
            response = self.session.post(url=all_tickers_get_url, headers=headers, json=payload)
            response.raise_for_status()
            data_response = response.json()
            self.all_tickers = [item["s"] for item in data_response["data"]]
        console.print("[bold green]✔ Получены тикеры[/bold green]")

    # -----------------------
    # Получить корреляцию тикера
    # -----------------------
    def get_correlation(self):
        ticker = self.driver.select(".valueValue-l31H9iuA.apply-common-tooltip").text
        correlation = float(
            self.driver.select(".sources-l31H9iuA .valueValue-l31H9iuA").text.replace(",", ".").replace("−", "-")
        )
        self.tickers_correlations[ticker] = correlation
    
    # -----------------------
    # Сбор всех корреляций
    # -----------------------
    def collect_correlations(self):
        with console.status("[yellow]Подготавливаем тикеры...[/yellow]"):
            self.driver.click("button[data-name='screener-dialog-button']")
            
            template_name_screener = self.driver.select("div[data-qa-id='screen-title']")
            if template_name_screener.text != "Фьючерсы Bybit":
                template_name_screener.click()
                self.driver.get_element_with_exact_text("Фьючерсы Bybit").click()

            total_tickers = len(self.all_tickers)
            tickers = self.driver.select_all("tbody tr")
        console.print("[bold green]✔ Список тикеров готов[/bold green]")

        console.print(f"[magenta]Найдено {total_tickers} тикеров...[/magenta]")

        progress = Progress(
            SpinnerColumn(),
            TextColumn("[cyan]{task.description}"),
            BarColumn(complete_style="green", finished_style="bright_green"),
            TextColumn("[bright_green]{task.completed}[/bright_green]/[yellow]{task.total}[/yellow]"),
            TextColumn("[bright_green]{task.percentage:>3.0f}%[/bright_green]"),
            TimeRemainingColumn()
        )
        
        with progress:
            task = progress.add_task("Собираем корреляции", total=total_tickers)
            i = 0
            while i < len(tickers):
                tickers[i].click()
                self.get_correlation()
                i += 1
                progress.update(task, advance=1)

                if i >= len(tickers) - 5:
                    tickers = self.driver.select_all("tbody tr")
                
                if len(self.tickers_correlations) == total_tickers:
                    break
        
        print()
        console.print(Panel.fit(
            "🎯 [bold white]Все корреляции собраны![/bold white] 🎯",
            border_style="bold green",
            style="on dark_green",
            title="[bold yellow]Завершено[/bold yellow]"
        ))
        print()
    
    # =======================
    # Вывод результатов
    # =======================

    # -----------------------
    # Получение настроек 
    # -----------------------
    def ask_user_settings(self):
        threshold = inquirer.select(
            message="Выберите порог корреляции:",
            choices=[f"{x/10:.1f}" for x in range(1, 11)],
            default="0.5",
            qmark="", amark=""
        ).execute()
        threshold = float(threshold)

        sort_order = inquirer.select(
            message="Сортировка:",
            choices=[
                {"name": "без сортировки", "value": None},
                {"name": "сначала низкая корреляция", "value": "asc"},
                {"name": "сначала высокая корреляция", "value": "desc"},
            ],
            default=None,
            qmark="", amark=""
        ).execute()

        return threshold, sort_order

    # -----------------------
    # Сортировка корреляций
    # -----------------------
    def sort_correlations(self, sort_order):
        if sort_order:
            self.tickers_correlations = dict(
                sorted(
                    self.tickers_correlations.items(),
                    key=lambda item: item[1],
                    reverse=(sort_order == "desc")
                )
            )

    # -----------------------
    # Сохранение результатов в файлы
    # -----------------------
    def save_results_to_files(self, threshold):
        file_name = f"Корреляция_{datetime.now():%d.%m.%y_%H-%M}"
        
        with console.status("[green]Сохраняем результаты в файлы...[/green]"):
            # TXT
            with open(f"{file_name}.txt", "w", encoding="utf-8") as f:
                for ticker, corr in self.tickers_correlations.items():
                    f.write(f"{ticker}: {corr}\n")
            # Excel
            wb = Workbook()
            ws = wb.active
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
            for ticker, corr in self.tickers_correlations.items():
                ws.append([ticker, corr])
                cell = ws.cell(row=ws.max_row, column=2)
                cell.fill = green_fill if corr <= threshold else red_fill
            wb.save(f"{file_name}.xlsx")
        
        console.print(Panel.fit(
            f"🎯 [bold white]Результаты успешно сохранены![/bold white] 🎯\n\n"
            f"[cyan]Файлы:[/cyan]\n"
            f"• {file_name}.txt\n"
            f"• {file_name}.xlsx",
            border_style="bold green"
        ))
    
    # -----------------------
    # Показ таблицы в терминале
    # -----------------------
    def show_results_table(self, threshold):
        show_table = inquirer.select(
            message="Показать таблицу корреляций?",
            choices=[{"name": "Да", "value": True}, {"name": "Нет", "value": False}],
            default=True,
            qmark="", amark=""
        ).execute()

        if not show_table:
            return

        show_all_rows = inquirer.select(
            message=f"Все или только с корреляцией ниже {threshold}?",
            choices=[{"name": f"Ниже {threshold}", "value": False}, {"name": "Все", "value": True}],
            default=False,
            qmark="", amark=""
        ).execute()

        if not show_all_rows:
            data = {ticker: corr for ticker, corr in self.tickers_correlations.items() if corr <= threshold}
        else:
            data = self.tickers_correlations

        table = Table(title=f"\nКорреляции тикеров ({len(data)})", show_lines=True)
        table.add_column("Тикер", justify="center", style="cyan", no_wrap=True)
        table.add_column("Корреляция", justify="center")
        
        for ticker, corr in data.items():
            table.add_row(ticker, f"[green]{corr}[/green]")

        console.print(table)

    # -----------------------
    # Добавление тикеров с низкой корреляцией в список TradingView
    # -----------------------
    def add_low_corr_tickers_to_tradingview(self, threshold):
        add_to_watchlist = inquirer.select(
            message=f"Добавить тикеры с корреляцией ниже {threshold} в список TradingView?",
            choices=[
                {"name": "Да", "value": True}, 
                {"name": "Нет", "value": False}
            ],
            default=True,
            qmark="", amark=""
        ).execute()

        if not add_to_watchlist:
            return

        low_corr_tickers = [ticker for ticker, corr in self.tickers_correlations.items() if corr <= threshold]
        console.print(f"[cyan]Найдено {len(low_corr_tickers)} тикеров с корреляцией ниже {threshold}[/cyan]")

        batch_size = 30
        total = len(low_corr_tickers)

        for i in range(0, total, batch_size):
            ticker_batch = low_corr_tickers[i:i + batch_size]
            start = i + 1
            end = i + len(ticker_batch)

            with console.status(f"[cyan]Добавляем {start}-{end} из {len(low_corr_tickers)} тикеров в список TradingView...[/cyan]", spinner="line"):
                self.add_tickers_to_list(ticker_batch)
            
            if i + batch_size >= total:
                console.print(f"[bold green]Все {total} тикеров были успешно добавлены в список![/bold green]")
                break
            
            console.print(f"[yellow]Тикеры {start}-{end} добавлены в список.[/yellow]")

            action = inquirer.select(
                message="Продолжить добавление следующих тикеров?",
                choices=[
                    {"name": "Продолжить", "value": "continue"}, 
                    {"name": "Пропустить оставшиеся", "value": "skip"}
                ],
                default="continue",
                qmark="", amark=""
            ).execute()

            if action == "skip":
                console.print("[red]Добавление тикеров остановлено пользователем[/red]")
                break

        input("\nНажмите Enter для завершения...")
        self.session.close()

    # -----------------------
    # Основная функция display_results (только оркестрация)
    # -----------------------
    def display_results(self):
        console.print(f"[cyan]{'═'*25}[/cyan] [bold white]Вывод результатов[/bold white] [cyan]{'═'*25}[/cyan]")
        
        threshold, sort_order = self.ask_user_settings()
        self.sort_correlations(sort_order)
        self.save_results_to_files(threshold)
        self.show_results_table(threshold)
        self.add_low_corr_tickers_to_watchlist(threshold)

    # -----------------------
    # Pipeline: корреляции
    # -----------------------
    def get_correlations(self):
        self.open_tradingview()
        self.get_all_tickers()
        self.activate_corr_indicator()
        self.collect_correlations()
        self.display_results()

        return self.tickers_correlations
